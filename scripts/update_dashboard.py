#!/usr/bin/env python3
"""
Atualiza o dashboard de Fluxo de Caixa (Business Connection) a partir da
planilha Excel hospedada no OneDrive, sem qualquer intervenção manual.
"""
import base64
import datetime as dt
import json
import os
import sys
import urllib.request
import urllib.error

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(HERE)
XLSX_PATH = os.path.join(REPO_ROOT, "_planilha_baixada.xlsx")
CHARTJS_PATH = os.path.join(HERE, "chart.umd.min.js")
OUTPUT_PATH = os.path.join(REPO_ROOT, "index.html")

COLORS = {
    "Mercury": "#33d19b",
    "Inter": "#ffb648",
    "Stripe": "#4f8cff",
    "Hotmart": "#ff6b6b",
    "IA For Business": "#a78bfa",
}


def _share_token(share_url: str) -> str:
    b64 = base64.urlsafe_b64encode(share_url.encode("utf-8")).decode("utf-8").rstrip("=")
    return "u!" + b64


def _candidate_urls(share_url: str):
    token = _share_token(share_url)
    return [
        f"https://graph.microsoft.com/v1.0/shares/{token}/driveItem/content",
        f"https://api.onedrive.com/v1.0/shares/{token}/root/content",
    ]


def download_spreadsheet(share_url: str, dest_path: str) -> None:
    last_error = None
    for api_url in _candidate_urls(share_url):
        req = urllib.request.Request(api_url, headers={"User-Agent": "dashboard-updater/1.0"})
        try:
            with urllib.request.urlopen(req, timeout=60) as resp, open(dest_path, "wb") as f:
                f.write(resp.read())
            return
        except urllib.error.HTTPError as e:
            last_error = e
            print(f"  tentativa via {api_url.split('/')[2]} falhou: HTTP {e.code}")
            continue
    raise SystemExit(
        "Falha ao baixar a planilha do OneDrive "
        f"(HTTP {last_error.code if last_error else '??'}). O link de "
        "compartilhamento precisa estar configurado como 'Qualquer pessoa "
        "com o link' com permissão de visualização/download. Verifique "
        "ONEDRIVE_SHARE_URL. Se o erro persistir em ambos os endpoints, a "
        "conta pode estar bloqueando acesso anônimo programático — nesse "
        "caso é necessário usar autenticação via Microsoft Graph (Azure AD)."
    ) from last_error


def load_workbook(path):
    return openpyxl.load_workbook(path, data_only=True)


def parse_fluxo_de_caixa(ws, today: dt.date):
    accounts_raw = {
        "Mercury": ws.cell(row=6, column=4).value,
        "Inter": ws.cell(row=6, column=5).value,
        "Stripe": ws.cell(row=6, column=6).value,
        "Hotmart": ws.cell(row=6, column=7).value,
        "IA For Business": ws.cell(row=6, column=9).value,
    }
    accounts = [
        {"label": k, "value": round(float(v or 0), 2), "color": COLORS[k]}
        for k, v in accounts_raw.items()
    ]
    accounts.sort(key=lambda a: a["value"], reverse=True)

    daily_rows = []
    for r in range(11, 195):
        date_val = ws.cell(row=r, column=2).value
        if not isinstance(date_val, dt.datetime):
            continue
        daily_rows.append(
            {
                "date": date_val.date(),
                "entrada_p": float(ws.cell(row=r, column=3).value or 0),
                "saida_p": float(ws.cell(row=r, column=4).value or 0),
                "saldo_p": float(ws.cell(row=r, column=5).value or 0),
                "entrada_r": float(ws.cell(row=r, column=8).value or 0),
                "saida_r": float(ws.cell(row=r, column=9).value or 0),
                "saldo_r": float(ws.cell(row=r, column=10).value or 0),
            }
        )

    today_row = next((row for row in daily_rows if row["date"] == today), None)
    if today_row is None:
        past = [row for row in daily_rows if row["date"] <= today]
        today_row = past[-1] if past else daily_rows[0]

    return accounts, daily_rows, today_row


def parse_contas_a_pagar(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        data_prev = ws.cell(row=r, column=1).value
        nome = ws.cell(row=r, column=2).value
        valor = ws.cell(row=r, column=3).value
        categoria = ws.cell(row=r, column=6).value
        if not isinstance(data_prev, dt.datetime) or valor in (None, ""):
            continue
        try:
            valor = float(valor)
        except (TypeError, ValueError):
            continue
        rows.append({"date": data_prev.date(), "nome": nome or "—", "valor": valor, "categoria": categoria or "—"})
    return rows


def parse_contas_a_receber(ws):
    rows = []
    for r in range(3, ws.max_row + 1):
        cliente = ws.cell(row=r, column=1).value
        previsao = ws.cell(row=r, column=3).value
        valor_liquido = ws.cell(row=r, column=7).value
        if not isinstance(previsao, dt.datetime) or valor_liquido in (None, ""):
            continue
        try:
            valor_liquido = float(valor_liquido)
        except (TypeError, ValueError):
            continue
        rows.append({"date": previsao.date(), "cliente": (cliente or "—").strip(), "valor": valor_liquido})
    return rows


def end_of_month(d: dt.date) -> dt.date:
    if d.month == 12:
        return d.replace(day=31)
    next_month = d.replace(day=1, month=d.month + 1)
    return next_month - dt.timedelta(days=1)


def end_of_next_month(d: dt.date) -> dt.date:
    first_this = d.replace(day=1)
    if first_this.month == 12:
        first_next = first_this.replace(year=first_this.year + 1, month=1)
    else:
        first_next = first_this.replace(month=first_this.month + 1)
    return end_of_month(first_next)


def find_row(daily_rows, target_date):
    for row in daily_rows:
        if row["date"] == target_date:
            return row
    prior = [row for row in daily_rows if row["date"] <= target_date]
    return prior[-1] if prior else None


def build_dashboard_data(daily_rows, today_row, accounts, pagar_rows, receber_rows):
    today = today_row["date"]
    fim_mes = end_of_month(today)
    fim_prox_mes = end_of_next_month(today)

    proj = [row for row in daily_rows if today < row["date"] <= fim_mes]

    saldo_15d_date = today + dt.timedelta(days=15)
    saldo_15d_row = find_row(daily_rows, saldo_15d_date)

    fim_mes_row = find_row(daily_rows, fim_mes)
    fim_prox_mes_row = find_row(daily_rows, fim_prox_mes)

    if proj:
        low_point = min(proj, key=lambda r: r["saldo_p"])
    else:
        low_point = today_row

    entradas_periodo = sum(r["entrada_p"] for r in proj)
    saidas_periodo = sum(r["saida_p"] for r in proj)

    pagamentos_no_dia = [p for p in pagar_rows if p["date"] == low_point["date"]]
    top_categoria = None
    if pagamentos_no_dia:
        by_cat = {}
        for p in pagamentos_no_dia:
            by_cat[p["categoria"]] = by_cat.get(p["categoria"], 0) + p["valor"]
        top_categoria = max(by_cat.items(), key=lambda kv: kv[1])

    janela_fim = today + dt.timedelta(days=15)
    recebimentos_proximos = sorted(
        [r for r in receber_rows if today < r["date"] <= janela_fim],
        key=lambda r: (-r["valor"]),
    )[:15]
    pagamentos_proximos = sorted(
        [p for p in pagar_rows if today < p["date"] <= janela_fim],
        key=lambda p: (-p["valor"]),
    )[:15]

    return {
        "today": today,
        "today_saldo": today_row["saldo_r"],
        "fim_mes": fim_mes,
        "fim_mes_saldo": fim_mes_row["saldo_p"] if fim_mes_row else None,
        "saldo_15d_date": saldo_15d_row["date"] if saldo_15d_row else None,
        "saldo_15d": saldo_15d_row["saldo_p"] if saldo_15d_row else None,
        "fim_prox_mes": fim_prox_mes,
        "fim_prox_mes_saldo": fim_prox_mes_row["saldo_p"] if fim_prox_mes_row else None,
        "low_point": low_point,
        "top_categoria": top_categoria,
        "entradas_periodo": entradas_periodo,
        "saidas_periodo": saidas_periodo,
        "proj": proj,
        "accounts": accounts,
        "recebimentos_proximos": recebimentos_proximos,
        "pagamentos_proximos": pagamentos_proximos,
        "janela_fim": janela_fim,
    }


def fmt_money(v):
    v = round(float(v or 0), 2)
    neg = v < 0
    v = abs(v)
    s = f"{v:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    return f"-$ {s}" if neg else f"$ {s}"


def fmt_ddmm(d: dt.date) -> str:
    return d.strftime("%d/%m")


def build_html(data: dict) -> str:
    today = data["today"]
    today_saldo = data["today_saldo"]
    fim_mes = data["fim_mes"]
    fim_mes_saldo = data["fim_mes_saldo"]
    saldo_15d = data["saldo_15d"]
    saldo_15d_date = data["saldo_15d_date"]
    fim_prox_mes = data["fim_prox_mes"]
    fim_prox_mes_saldo = data["fim_prox_mes_saldo"]
    low_point = data["low_point"]
    top_categoria = data["top_categoria"]
    entradas_periodo = data["entradas_periodo"]
    saidas_periodo = data["saidas_periodo"]
    proj = data["proj"]
    accounts = data["accounts"]
    recebimentos = data["recebimentos_proximos"]
    pagamentos = data["pagamentos_proximos"]
    janela_fim = data["janela_fim"]

    variacao = (fim_mes_saldo - today_saldo) if fim_mes_saldo is not None else 0
    variacao_cls = "pos" if variacao >= 0 else "neg"
    variacao_sinal = "+" if variacao >= 0 else ""

    dip_cls = "amber" if low_point["saldo_p"] >= 0 else "red"
    dip_cause = (
        f" — principal fator: {top_categoria[0].title()} ({fmt_money(top_categoria[1])})"
        if top_categoria
        else ""
    )

    kpi_html = f"""
<div class="kpis">
  <div class="kpi">
    <div class="label">Saldo do Dia (hoje)</div>
    <div class="value">{fmt_money(today_saldo)}</div>
    <div class="note">{fmt_ddmm(today)}/{today.year} · realizado</div>
  </div>
  <div class="kpi">
    <div class="label">Saldo em 15 dias</div>
    <div class="value {'pos' if (saldo_15d or 0) >= today_saldo else 'neg'}">{fmt_money(saldo_15d) if saldo_15d is not None else '—'}</div>
    <div class="note">{fmt_ddmm(saldo_15d_date) if saldo_15d_date else '—'}</div>
  </div>
  <div class="kpi">
    <div class="label">Saldo em {fmt_ddmm(fim_mes)}</div>
    <div class="value {'pos' if (fim_mes_saldo or 0) >= today_saldo else 'neg'}">{fmt_money(fim_mes_saldo) if fim_mes_saldo is not None else '—'}</div>
    <div class="note">fim do mês</div>
  </div>
  <div class="kpi">
    <div class="label">Ponto Mais Baixo</div>
    <div class="value {dip_cls}-txt">{fmt_money(low_point['saldo_p'])}</div>
    <div class="note">{fmt_ddmm(low_point['date'])}</div>
  </div>
  <div class="kpi">
    <div class="label">Variação Líquida (até {fmt_ddmm(fim_mes)})</div>
    <div class="value {variacao_cls}">{variacao_sinal}{fmt_money(variacao)}</div>
    <div class="note">{fmt_money(entradas_periodo)} entradas − {fmt_money(saidas_periodo)} saídas</div>
  </div>
</div>
"""

    prox_mes_html = f"""
<div class="card" style="margin-bottom:18px; display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; gap:10px;">
  <div>
    <h3 style="margin-bottom:2px;">Saldo Projetado em {fmt_ddmm(fim_prox_mes)}</h3>
    <div class="desc" style="margin-bottom:0;">Último dia do mês subsequente</div>
  </div>
  <div style="text-align:right;">
    <div class="value {'pos' if (fim_prox_mes_saldo or 0) >= today_saldo else 'neg'}" style="font-size:22px; font-weight:700;">{fmt_money(fim_prox_mes_saldo) if fim_prox_mes_saldo is not None else '—'}</div>
    <div class="note">{'+' if (fim_prox_mes_saldo or 0) >= today_saldo else ''}{fmt_money((fim_prox_mes_saldo or 0) - today_saldo)} vs. saldo de hoje</div>
  </div>
</div>
"""

    story_html = f"""
<div class="story">
  <div class="icon">📅</div>
  <div>
    <h3 class="{dip_cls}">Ponto mais apertado do mês: {fmt_ddmm(low_point['date'])}</h3>
    <p>Com base na planilha atualizada, o saldo projetado mais baixo do período é <strong>{fmt_money(low_point['saldo_p'])}</strong>, em <strong>{fmt_ddmm(low_point['date'])}</strong>{dip_cause}.</p>
    <p>Considerando entradas e saídas previstas até {fmt_ddmm(fim_mes)}, o saldo termina o mês em <strong>{fmt_money(fim_mes_saldo) if fim_mes_saldo is not None else '—'}</strong>{' — ' + fmt_money(variacao) + ' acima do saldo de hoje.' if variacao >= 0 else ', uma redução de ' + fmt_money(abs(variacao)) + ' em relação ao saldo de hoje.'}</p>
  </div>
</div>
"""

    def recebimentos_rows():
        out = []
        for r in recebimentos:
            out.append(
                f"<tr><td>{r['cliente']}</td><td>{fmt_ddmm(r['date'])}</td>"
                f"<td class=\"pos-cell\">{fmt_money(r['valor'])}</td></tr>"
            )
        return "\n".join(out) if out else "<tr><td colspan=3>Sem recebimentos previstos no período.</td></tr>"

    def pagamentos_rows():
        out = []
        for p in pagamentos:
            out.append(
                f"<tr><td>{p['nome']}</td><td>{fmt_ddmm(p['date'])}</td>"
                f"<td class=\"neg-cell\">{fmt_money(p['valor'])}</td></tr>"
            )
        return "\n".join(out) if out else "<tr><td colspan=3>Sem pagamentos previstos no período.</td></tr>"

    agenda_html = f"""
<div class="section-title">Agenda de recebimentos e pagamentos — próximos 15 dias ({fmt_ddmm(today)} a {fmt_ddmm(janela_fim)})</div>
<div class="grid" style="display:grid; grid-template-columns:1fr 1fr; gap:18px;">
  <div class="card">
    <h3>Principais Recebimentos</h3>
    <div class="desc">Valor líquido previsto, por data</div>
    <table>
      <thead><tr><th>Cliente</th><th>Data</th><th>Valor</th></tr></thead>
      <tbody>
        {recebimentos_rows()}
      </tbody>
    </table>
  </div>
  <div class="card">
    <h3>Principais Pagamentos</h3>
    <div class="desc">Itens de maior valor no período, por data</div>
    <table>
      <thead><tr><th>Pagamento</th><th>Data</th><th>Valor</th></tr></thead>
      <tbody>
        {pagamentos_rows()}
      </tbody>
    </table>
  </div>
</div>
"""

    chartjs_lib = open(CHARTJS_PATH, encoding="utf-8").read()

    proj_json = json.dumps(
        [
            {
                "date": p["date"].isoformat(),
                "entrada": round(p["entrada_p"], 2),
                "saida": round(p["saida_p"], 2),
                "net": round(p["entrada_p"] - p["saida_p"], 2),
                "saldo": round(p["saldo_p"], 2),
            }
            for p in proj
        ]
    )
    accounts_json = json.dumps(accounts)
    dip_date_iso = low_point["date"].isoformat()

    script_html = f"""
<script>
{chartjs_lib}
</script>
<script>
const today = {{ date: '{today.isoformat()}', saldo: {round(today_saldo, 2)} }};
const proj = {proj_json};
const accounts = {accounts_json};
const accountsTotal = accounts.reduce((s,a)=>s+a.value,0);
const dipDate = '{dip_date_iso}';

function fmt(v) {{
  const neg = v < 0; v = Math.abs(v);
  const s = '$ ' + v.toLocaleString('pt-BR', {{minimumFractionDigits:2, maximumFractionDigits:2}});
  return neg ? '-' + s : s;
}}
function fmtDate(iso) {{
  const [y,m,d] = iso.split('-');
  return d+'/'+m;
}}

const labels = [fmtDate(today.date), ...proj.map(p=>fmtDate(p.date))];
const entradaData = [null, ...proj.map(p=>p.entrada)];
const saidaData = [null, ...proj.map(p=>-p.saida)];
const saldoData = [today.saldo, ...proj.map(p=>p.saldo)];

new Chart(document.getElementById('chartFlow'), {{
  data: {{
    labels: labels,
    datasets: [
      {{ type:'bar', label:'Entradas', data: entradaData, backgroundColor:'#33d19b', borderRadius:3, yAxisID:'y1', order:2 }},
      {{ type:'bar', label:'Saídas', data: saidaData, backgroundColor:'#ff6b6b', borderRadius:3, yAxisID:'y1', order:2 }},
      {{ type:'line', label:'Saldo projetado', data: saldoData, borderColor:'#4f8cff', backgroundColor:'rgba(79,140,255,.15)', borderWidth:2.5, tension:.3, fill:true, pointRadius:3, pointBackgroundColor:'#4f8cff', yAxisID:'y2', order:1 }}
    ]
  }},
  options: {{
    responsive:true, maintainAspectRatio:false,
    interaction:{{ mode:'index', intersect:false }},
    plugins:{{
      legend:{{ labels:{{ color:'#98a2b8', usePointStyle:true, boxWidth:8 }} }},
      tooltip:{{ callbacks:{{ label: function(ctx){{
        if (ctx.dataset.type === 'bar') return ctx.dataset.label + ': ' + fmt(Math.abs(ctx.parsed.y || 0));
        return ctx.dataset.label + ': ' + fmt(ctx.parsed.y);
      }} }} }}
    }},
    scales:{{
      x:{{ ticks:{{ color:'#98a2b8', maxRotation:0, autoSkip:true, autoSkipPadding:6 }}, grid:{{ color:'#2a334a' }} }},
      y1:{{ position:'left', ticks:{{ color:'#98a2b8', callback:v=>fmt(Math.abs(v)) }}, grid:{{ color:'#2a334a' }}, title:{{display:true, text:'Entrada / Saída', color:'#98a2b8'}} }},
      y2:{{ position:'right', ticks:{{ color:'#98a2b8', callback:v=>fmt(v) }}, grid:{{ drawOnChartArea:false }}, title:{{display:true, text:'Saldo', color:'#98a2b8'}} }}
    }}
  }}
}});

new Chart(document.getElementById('chartAccounts'), {{
  type:'doughnut',
  data:{{ labels: accounts.map(a=>a.label), datasets:[{{ data: accounts.map(a=>a.value), backgroundColor: accounts.map(a=>a.color), borderColor:'#171f30', borderWidth:3 }}] }},
  options:{{
    responsive:true, maintainAspectRatio:false, cutout:'62%',
    plugins:{{ legend:{{ display:false }}, tooltip:{{ callbacks:{{ label: ctx => ctx.label + ': ' + fmt(ctx.parsed) }} }} }}
  }}
}});

const listEl = document.getElementById('accountsList');
accounts.forEach(a=>{{
  const pct = (a.value/accountsTotal*100).toFixed(1);
  const row = document.createElement('div');
  row.className = 'dexp-row';
  row.innerHTML = `
    <div class="dexp-top">
      <span class="dexp-dot" style="background:${{a.color}}"></span>
      <span class="dexp-label">${{a.label}}</span>
      <span class="dexp-val">${{fmt(a.value)}}</span>
      <span class="dexp-pct">${{pct}}%</span>
    </div>
    <div class="dexp-bar-bg"><div class="dexp-bar" style="width:${{pct}}%; background:${{a.color}}"></div></div>
  `;
  listEl.appendChild(row);
}});
const totalRow = document.createElement('div');
totalRow.className = 'dexp-row hl';
totalRow.innerHTML = `<div class="dexp-top"><span class="dexp-dot" style="background:transparent"></span><span class="dexp-label">Total</span><span class="dexp-val">${{fmt(accountsTotal)}}</span><span class="dexp-pct">100%</span></div>`;
listEl.appendChild(totalRow);

const tbody = document.querySelector('#dailyTable tbody');
const todayRow = document.createElement('tr');
todayRow.className = 'today';
todayRow.innerHTML = `<td>${{fmtDate(today.date)}} (hoje)</td><td>—</td><td>—</td><td>—</td><td>${{fmt(today.saldo)}}</td>`;
tbody.appendChild(todayRow);

proj.forEach(p=>{{
  const tr = document.createElement('tr');
  if (p.date === dipDate) tr.className = 'dip';
  tr.innerHTML = `
    <td>${{fmtDate(p.date)}}</td>
    <td class="pos-cell">${{fmt(p.entrada)}}</td>
    <td class="neg-cell">${{fmt(p.saida)}}</td>
    <td class="${{p.net>=0?'pos-cell':'neg-cell'}}">${{p.net>=0?'+':''}}${{fmt(p.net)}}</td>
    <td>${{fmt(p.saldo)}}</td>
  `;
  tbody.appendChild(tr);
}});
</script>
"""

    now_str = dt.datetime.now().strftime("%d/%m/%Y %H:%M")

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Fluxo de Caixa — Business Connection</title>
<style>
:root{{
  --bg:#0f1420; --panel:#161d2e; --panel2:#1c2438; --border:#2a334a;
  --text:#eef1f7; --muted:#98a2b8;
  --green:#33d19b; --red:#ff6b6b; --blue:#4f8cff; --amber:#ffb648; --purple:#a78bfa;
}}
*{{box-sizing:border-box;}}
body{{
  margin:0; background:linear-gradient(180deg,#0b0f18,#0f1420 300px); color:var(--text);
  font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;
  padding:32px 40px 60px;
}}
h1{{font-size:24px; margin:0 0 4px;}}
.subtitle{{color:var(--muted); font-size:13.5px; margin-bottom:26px;}}

.kpis{{display:grid; grid-template-columns:repeat(5,1fr); gap:14px; margin-bottom:22px;}}
.kpi{{background:var(--panel); border:1px solid var(--border); border-radius:12px; padding:16px 18px;}}
.kpi .label{{font-size:12px; color:var(--muted); text-transform:uppercase; letter-spacing:.04em;}}
.kpi .value{{font-size:21px; font-weight:700; margin-top:6px; font-variant-numeric:tabular-nums;}}
.kpi .note{{font-size:12px; margin-top:4px; color:var(--muted);}}
.pos{{color:var(--green);}} .neg{{color:var(--red);}} .amber-txt{{color:var(--amber);}} .red-txt{{color:var(--red);}}

.story{{
  background:linear-gradient(135deg,#1c2438,#20263a); border:1px solid #3a4566; border-radius:12px;
  padding:20px 22px; margin-bottom:18px; display:flex; gap:16px; align-items:flex-start;
}}
.story .icon{{font-size:26px;}}
.story h3{{margin:0 0 6px; font-size:16px;}}
.story h3.amber{{color:var(--amber);}} .story h3.red{{color:var(--red);}} .story h3.green{{color:var(--green);}}
.story p{{margin:0; color:#d3d9e6; font-size:13.5px; line-height:1.6;}}
.story p+p{{margin-top:8px;}}

.featured{{
  background:linear-gradient(160deg,#1c2438,#171f30);
  border:1px solid #3a4566; border-radius:16px; padding:26px 28px; margin-bottom:20px;
  box-shadow:0 0 0 1px rgba(255,182,72,.08), 0 12px 32px rgba(0,0,0,.35);
}}
.featured-head{{display:flex; align-items:baseline; justify-content:space-between; gap:12px; flex-wrap:wrap; margin-bottom:16px;}}
.featured-head h2{{margin:0; font-size:17px;}}
.featured-head .badge{{background:rgba(255,182,72,.15); color:var(--amber); font-size:12px; font-weight:700; padding:4px 10px; border-radius:20px;}}
.featured-body{{display:grid; grid-template-columns:1.1fr 1.4fr; gap:28px; align-items:center;}}

.dexp-list{{display:flex; flex-direction:column; gap:12px;}}
.dexp-row.hl .dexp-label, .dexp-row.hl .dexp-val{{color:var(--amber); font-weight:700;}}
.dexp-top{{display:flex; align-items:center; gap:8px; font-size:13px; margin-bottom:5px;}}
.dexp-dot{{width:9px; height:9px; border-radius:50%; flex-shrink:0;}}
.dexp-label{{flex:1; color:#dfe4ee;}}
.dexp-val{{color:#c7cee0; font-variant-numeric:tabular-nums;}}
.dexp-pct{{width:44px; text-align:right; color:var(--muted); font-variant-numeric:tabular-nums;}}
.dexp-bar-bg{{height:7px; background:rgba(255,255,255,.06); border-radius:6px; overflow:hidden;}}
.dexp-bar{{height:100%; border-radius:6px;}}

.chart-wrap{{position:relative; height:280px; width:100%;}}
.chart-wrap.tall{{height:340px;}}

.card{{background:var(--panel); border:1px solid var(--border); border-radius:12px; padding:18px 20px;}}
.card h3{{margin:0 0 2px; font-size:15px;}}
.card .desc{{color:var(--muted); font-size:12.5px; margin-bottom:12px;}}

table{{width:100%; border-collapse:collapse; font-size:13px;}}
th,td{{padding:7px 9px; text-align:right; border-bottom:1px solid var(--border);}}
th:first-child, td:first-child{{text-align:left; color:var(--muted);}}
tr.total td{{font-weight:700; border-top:1px solid #3a4566;}}
tr.today td{{background:rgba(79,140,255,.08);}}
tr.dip td{{background:rgba(255,107,107,.08);}}
td.neg-cell{{color:var(--red);}} td.pos-cell{{color:var(--green);}}

.section-title{{font-size:14px; color:var(--muted); text-transform:uppercase; letter-spacing:.05em; margin:26px 0 12px;}}
.updated-badge{{display:inline-block; background:rgba(79,140,255,.15); color:var(--blue); font-size:11.5px; font-weight:700; padding:3px 9px; border-radius:14px; margin-left:8px; vertical-align:middle;}}
@media (max-width:1100px){{ .kpis{{grid-template-columns:repeat(2,1fr);}} .grid,.grid-32,.featured-body{{grid-template-columns:1fr;}} }}
</style>
</head>
<body>

<h1>Fluxo de Caixa — Business Connection <span class="updated-badge">atualizado {now_str}</span></h1>
<div class="subtitle">Saldo do dia e projeção até o fim do mês · dados de {fmt_ddmm(today)}/{today.year} a {fmt_ddmm(fim_mes)}/{fim_mes.year} · valores em $ · atualizado automaticamente todo dia a partir da planilha do OneDrive</div>

{kpi_html}
{prox_mes_html}
{story_html}

<div class="section-title">Projeção diária — {fmt_ddmm(today)} a {fmt_ddmm(fim_mes)}</div>
<div class="card" style="margin-bottom:20px;">
  <h3>Saldo projetado dia a dia</h3>
  <div class="desc">Barras = entrada e saída previstas no dia · Linha = saldo acumulado projetado</div>
  <div class="chart-wrap tall"><canvas id="chartFlow"></canvas></div>
</div>

<div class="featured">
  <div class="featured-head">
    <h2>Saldo por Conta (Business Connection)</h2>
    <span class="badge">Snapshot mais recente</span>
  </div>
  <div class="featured-body">
    <div class="chart-wrap"><canvas id="chartAccounts"></canvas></div>
    <div class="dexp-list" id="accountsList"></div>
  </div>
</div>

{agenda_html}

<div class="section-title">Detalhamento diário — apêndice</div>
<div class="card">
  <h3>Saldo dia a dia (previsto)</h3>
  <div class="desc">Ponto de partida: saldo real de hoje ({fmt_money(today_saldo)}) + movimentação prevista de cada dia até {fmt_ddmm(fim_mes)}</div>
  <table id="dailyTable">
    <thead><tr><th>Data</th><th>Entradas</th><th>Saídas</th><th>Líquido do dia</th><th>Saldo Projetado</th></tr></thead>
    <tbody></tbody>
  </table>
</div>

{script_html}
</body>
</html>
"""


def main():
    share_url = os.environ.get("ONEDRIVE_SHARE_URL")
    if not share_url:
        print("ERRO: defina a variável de ambiente ONEDRIVE_SHARE_URL", file=sys.stderr)
        sys.exit(1)

    if os.environ.get("DASHBOARD_TODAY"):
        today = dt.date.fromisoformat(os.environ["DASHBOARD_TODAY"])
    else:
        today = dt.date.today()

    print("Baixando planilha do OneDrive...")
    download_spreadsheet(share_url, XLSX_PATH)

    print("Lendo planilha...")
    wb = load_workbook(XLSX_PATH)
    accounts, daily_rows, today_row = parse_fluxo_de_caixa(wb["Fluxo de Caixa"], today)
    pagar_rows = parse_contas_a_pagar(wb["Contas a Pagar"])
    receber_rows = parse_contas_a_receber(wb["Contas a Receber"])

    print("Calculando KPIs...")
    data = build_dashboard_data(daily_rows, today_row, accounts, pagar_rows, receber_rows)

    print("Gerando HTML...")
    html = build_html(data)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    os.remove(XLSX_PATH)
    print(f"Dashboard atualizado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
